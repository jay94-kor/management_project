import streamlit as st
import pandas as pd
import sqlite3
from openpyxl import load_workbook
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# 데이터베이스 연결 생성 및 테이블 생성 함수
def create_connection(db_file):
    """Create a database connection to the SQLite database specified by db_file"""
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except sqlite3.Error as e:
        print(e)
    return conn

def create_table(conn, create_table_sql):
    """Create a table from the create_table_sql statement"""
    try:
        c = conn.cursor()
        c.execute(create_table_sql)
    except sqlite3.Error as e:
        st.error(f"Error creating table: {e}")

def initialize_database():
    database = "project_budget_management.db"

    sql_create_projects_table = """
    CREATE TABLE IF NOT EXISTS projects (
        id integer PRIMARY KEY,
        name text NOT NULL,
        manager text,
        client text,
        start_date text,
        end_date text,
        first_draft_date text,
        event_location text,
        final_draft_date text
    );"""

    sql_create_budget_items_table = """
    CREATE TABLE IF NOT EXISTS budget_items (
        id integer PRIMARY KEY,
        project_id integer NOT NULL,
        category text,
        subcategory text,
        item text,
        quantity integer,
        unit text,
        days integer,
        unit_spec1 text,
        times integer,
        unit_spec2 text,
        price real,
        allocated_budget real,
        actual_cost real,
        FOREIGN KEY (project_id) REFERENCES projects (id)
    );"""

    sql_create_logs_table = """
    CREATE TABLE IF NOT EXISTS logs (
        id integer PRIMARY KEY,
        project_id integer NOT NULL,
        action text NOT NULL,
        timestamp text NOT NULL,
        details text,
        FOREIGN KEY (project_id) REFERENCES projects (id)
    );"""

    sql_create_users_table = """
    CREATE TABLE IF NOT EXISTS users (
        id integer PRIMARY KEY,
        username text NOT NULL,
        password text NOT NULL,
        role text NOT NULL
    );"""

    conn = create_connection(database)

    if conn is not None:
        create_table(conn, sql_create_projects_table)
        create_table(conn, sql_create_budget_items_table)
        create_table(conn, sql_create_logs_table)
        create_table(conn, sql_create_users_table)
    else:
        print("Error! Cannot create the database connection.")

# 로그 기록 및 조회 함수
def log_action(conn, project_id, action, details):
    try:
        sql = ''' INSERT INTO logs(project_id, action, timestamp, details)
                  VALUES(?,?,?,?) '''
        cur = conn.cursor()
        cur.execute(sql, (project_id, action, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), details))
        conn.commit()
    except sqlite3.Error as e:
        st.error(f"Error logging action: {e}")

def get_logs(conn):
    cur = conn.cursor()
    cur.execute("SELECT * FROM logs")
    rows = cur.fetchall()
    return rows

# 예산 조정 함수
def execute_query(conn, query, params=()):
    try:
        cur = conn.cursor()
        cur.execute(query, params)
        conn.commit()
        return cur
    except sqlite3.Error as e:
        st.error(f"Database error: {e}")
        return None

def transfer_budget_item(conn, from_item_id, to_item_id, amount):
    from_item = execute_query(conn, "SELECT allocated_budget, actual_cost FROM budget_items WHERE id=?", (from_item_id,)).fetchone()
    to_item = execute_query(conn, "SELECT allocated_budget, actual_cost FROM budget_items WHERE id=?", (to_item_id,)).fetchone()

    if from_item and to_item:
        from_allocated_budget, from_actual_cost = from_item
        to_allocated_budget, to_actual_cost = to_item

        if from_allocated_budget - from_actual_cost >= amount:
            new_from_allocated_budget = from_allocated_budget - amount
            new_to_allocated_budget = to_allocated_budget + amount

            execute_query(conn, "UPDATE budget_items SET allocated_budget=? WHERE id=?", (new_from_allocated_budget, from_item_id))
            execute_query(conn, "UPDATE budget_items SET allocated_budget=? WHERE id=?", (new_to_allocated_budget, to_item_id))

            log_action(conn, None, f"Transferred {amount} from item {from_item_id} to item {to_item_id}")
            return True
    return False

def transfer_budget_project(conn, from_project_id, to_project_id, amount):
    cur = conn.cursor()

    cur.execute("SELECT SUM(allocated_budget - actual_cost) FROM budget_items WHERE project_id=?", (from_project_id,))
    from_project_budget = cur.fetchone()[0]
    cur.execute("SELECT SUM(allocated_budget - actual_cost) FROM budget_items WHERE project_id=?", (to_project_id,))
    to_project_budget = cur.fetchone()[0]

    if from_project_budget and to_project_budget:
        if from_project_budget >= amount:
            cur.execute("""
                UPDATE budget_items
                SET allocated_budget = allocated_budget - ?
                WHERE project_id = ? AND (allocated_budget - actual_cost) >= ?
                LIMIT 1
            """, (amount, from_project_id, amount))
            cur.execute("""
                UPDATE budget_items
                SET allocated_budget = allocated_budget + ?
                WHERE project_id = ?
                LIMIT 1
            """, (amount, to_project_id))

            log_action(conn, None, f"Transferred {amount} from project {from_project_id} to project {to_project_id}")
            conn.commit()
            return True
    return False

# 예산 변경 요청 및 승인/반려 함수
def request_budget_change(conn, project_id, item_id, new_amount):
    sql = ''' INSERT INTO logs(project_id, action, timestamp, details)
              VALUES(?,?,?,?) '''
    cur = conn.cursor()
    details = f"Request to change budget item {item_id} to {new_amount}"
    cur.execute(sql, (project_id, "request", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), details))
    conn.commit()

def approve_budget_change(conn, log_id):
    cur = conn.cursor()
    cur.execute("SELECT details FROM logs WHERE id=?", (log_id,))
    row = cur.fetchone()
    if row:
        details = row[0]
        parts = details.split()
        item_id = parts[4]
        new_amount = float(parts[-1])

        cur.execute("UPDATE budget_items SET allocated_budget=? WHERE id=?", (new_amount, item_id))
        cur.execute("UPDATE logs SET action='approved' WHERE id=?", (log_id,))
        conn.commit()
        log_action(conn, None, "approve", f"Approved budget change for item {item_id} to {new_amount}")

def reject_budget_change(conn, log_id, reason):
    cur = conn.cursor()
    cur.execute("UPDATE logs SET action='rejected', details=details || ' Rejected: ' || ? WHERE id=?", (reason, log_id))
    conn.commit()

# 대시보드 함수
def show_dashboard():
    st.header("Dashboard")
    conn = create_connection("project_budget_management.db")
    if conn is not None:
        projects_df = pd.read_sql_query("SELECT * FROM projects", conn)
        budget_items_df = pd.read_sql_query("SELECT * FROM budget_items", conn)
        
        st.subheader("Projects Overview")
        st.dataframe(projects_df)
        
        st.subheader("Budget Items Overview")
        st.dataframe(budget_items_df)
        
        total_allocated_budget = budget_items_df['allocated_budget'].sum()
        total_actual_cost = budget_items_df['actual_cost'].sum()
        
        st.metric("Total Allocated Budget", f"${total_allocated_budget:,.2f}")
        st.metric("Total Actual Cost", f"${total_actual_cost:,.2f}")
        st.metric("Remaining Budget", f"${(total_allocated_budget - total_actual_cost):,.2f}")
    else:
        st.error("Database connection failed!")

# 구글 시트 데이터 업로드 함수
def authenticate_google_sheets(json_file):
    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name(json_file, scope)
    client = gspread.authorize(creds)
    return client

def fetch_data_from_google_sheets(client, sheet_name):
    sheet = client.open(sheet_name).sheet1
    data = sheet.get_all_records()
    return pd.DataFrame(data)

def upload_data():
    st.header("Upload Data")
    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])
    if uploaded_file:
        wb = load_workbook(uploaded_file)
        sheet = wb['예산 배정 및 정산서']

        project_name = sheet['C3'].value
        author_department = sheet['C4'].value
        client = sheet['H3'].value
        first_draft_date = sheet['H4'].value
        event_location = sheet['M3'].value
        final_draft_date = sheet['M4'].value
        event_start_date = sheet['O3'].value
        event_end_date = sheet['O4'].value

        data = []
        for row in sheet.iter_rows(min_row=8, min_col=1, max_col=13, values_only=True):
            if row[0] is None:
                break
            data.append({
                'category': row[0],
                'subcategory': row[1],
                'item': row[2],
                'description': row[3],
                'quantity': row[4],
                'unit': row[5],
                'days': row[6],
                'unit_spec1': row[7],
                'times': row[8],
                'unit_spec2': row[9],
                'price': row[10],
                'allocated_budget': row[11]
            })

        for row in sheet.iter_rows(values_only=True):
            if row and row[0] == "제안가 (VAT 포함)":
                proposal_amount = row[12]
                break

        conn = create_connection("project_budget_management.db")
        if conn is not None:
            cur = conn.cursor()
            cur.execute('''INSERT INTO projects (name, manager, client, start_date, end_date, first_draft_date, event_location, final_draft_date)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                        (project_name, author_department, client, event_start_date, event_end_date, first_draft_date, event_location, final_draft_date))
            project_id = cur.lastrowid

            for item in data:
                cur.execute('''INSERT INTO budget_items (project_id, category, subcategory, item, quantity, unit, days, unit_spec1, times, unit_spec2, price, allocated_budget)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                            (project_id, item['category'], item['subcategory'], item['item'], item['quantity'],
                             item['unit'], item['days'], item['unit_spec1'], item['times'], item['unit_spec2'],
                             item['price'], item['allocated_budget']))

            conn.commit()
            st.success("Data successfully uploaded to the database!")
        else:
            st.error("Database connection failed!")

    st.header("Google Sheets Data")
    json_file = st.file_uploader("Upload your Google Sheets API credentials", type=["json"])
    sheet_name = st.text_input("Enter the Google Sheet name")
    if st.button("Fetch Data"):
        if json_file and sheet_name:
            client = authenticate_google_sheets(json_file)
            df = fetch_data_from_google_sheets(client, sheet_name)
            st.dataframe(df)
            conn = create_connection("project_budget_management.db")
            if conn is not None:
                df.to_sql('budget_items', conn, if_exists='append', index=False)
                st.success("Google Sheets data successfully uploaded to the database!")
            else:
                st.error("Database connection failed!")

# 로그인 함수
def login():
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        conn = create_connection("project_budget_management.db")
        if conn is not None:
            cur = conn.cursor()
            cur.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
            user = cur.fetchone()
            if user:
                st.session_state['logged_in'] = True
                st.success("Login successful")
            else:
                st.error("Incorrect username or password")
        else:
            st.error("Database connection failed")

# 메인 함수
def main():
    if 'logged_in' not in st.session_state:
        st.session_state['logged_in'] = False

    if not st.session_state['logged_in']:
        login()
    else:
        st.title("프로젝트 예산 관리 시스템")
        st.sidebar.title("Navigation")
        options = st.sidebar.radio("Go to", ["Dashboard", "Upload Data", "Manage Budget", "Logs"])

        if options == "Dashboard":
            show_dashboard()
        elif options == "Upload Data":
            upload_data()
        elif options == "Manage Budget":
            st.write("예산 관리 기능")
        elif options == "Logs":
            conn = create_connection("project_budget_management.db")
            logs = get_logs(conn)
            st.dataframe(logs)

if __name__ == '__main__':
    initialize_database()
    main()