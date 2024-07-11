import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from app.db import create_connection
import sqlite3

def upload_data():
    st.header("Upload Data")
    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])
    if uploaded_file:
        wb = load_workbook(uploaded_file)
        sheet = wb['예산 배정 및 정산서']

        # Extracting project details
        project_name = sheet['C3'].value
        author_department = sheet['C4'].value
        client = sheet['H3'].value
        first_draft_date = sheet['H4'].value
        event_location = sheet['M3'].value
        final_draft_date = sheet['M4'].value
        event_start_date = sheet['O3'].value
        event_end_date = sheet['O4'].value

        # Extracting budget items
        data = []
        for row in sheet.iter_rows(min_row=8, min_col=1, max_col=13, values_only=True):
            if row[0] is None:
                break
            data.append({
                'category': row[0],
                'subcategory': row[1],
                'item': row[2],
                'description': row[3],
                'quantity': row[4],
                'unit': row[5],
                'days': row[6],
                'unit_spec1': row[7],
                'times': row[8],
                'unit_spec2': row[9],
                'price': row[10],
                'allocated_budget': row[11]
            })

        # Extracting proposal amount
        for row in sheet.iter_rows(values_only=True):
            if row and row[0] == "제안가 (VAT 포함)":
                proposal_amount = row[12]
                break

        conn = create_connection("project_budget_management.db")
        if conn is not None:
            cur = conn.cursor()
            # Insert project details
            cur.execute('''INSERT INTO projects (name, manager, client, start_date, end_date)
                           VALUES (?, ?, ?, ?, ?)''',
                        (project_name, author_department, client, event_start_date, event_end_date))
            project_id = cur.lastrowid

            # Insert budget items
            for item in data:
                cur.execute('''INSERT INTO budget_items (project_id, category, subcategory, item, quantity, unit, price, allocated_budget)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                            (project_id, item['category'], item['subcategory'], item['item'], item['quantity'],
                             item['unit'], item['price'], item['allocated_budget']))

            conn.commit()
            st.success("Data successfully uploaded to the database!")
        else:
            st.error("Database connection failed!")

